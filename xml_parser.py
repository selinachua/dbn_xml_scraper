'''
# Created by:
# Selina Chua
# selina.a.chua@gmail.com
#
# This program scrapes the XML files downloaded from the government site
# and grabs all the wanted information into an excel spreadsheet. The schema
# and architecture of the XML files are declared in the .xsd files that
# can be found in the same download bundle found in the link below. 
#
# The program creates a single Excel spreadsheet for every .xml file it 
# scrapes. It assumes that the .xml files it needs to scrape can be found in
# the ./privatehealth-04-apr-2019 directory, relative to the current directory.
# The excel files that are then populated are stored in the ./results directory, 
# also relative to cur directory.
#
# This program also depend on the parse_funds.py file, which parses the funds.xml
# file. This program requires the information from that program in order to provide
# complete information.
#
# WEBSITE OF DOWNLOAD:
# https://data.gov.au/dataset/ds-dga-8ab10b1f-6eac-423c-abc5-bbffc31b216c/details?q=%20private
'''

from bs4 import BeautifulSoup
import sys, os, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from hosp import HospService
from policy import NewPolicy, OldPolicy
from general import GeneralService
from parse_funds import parse_funds_file, Fund
from constants import *


def get_tag_name(tag):
    '''
    Gets tag name from a str(tag).
    '''
    start_ind = str(tag).find('<') + 1
    end_ind   = str(tag).find('>') or str(tag).find(' ')
    name = tag[start_ind:end_ind]
    return name

def get_excess(soup):
    '''
    Get all the excess information required.
    '''
    excesses = soup.find_all("excesses")[0].findChildren()
    excess_str = ""
    for excess in excesses:
        start_ind = str(excess).find('<') + 1
        end_ind = str(excess).find('>')
        excess_name = str(excess)[start_ind:end_ind]
        if "waiver" in excess_name:
            excess_str += f"{excess_name} for {excess.get_text()}"
            continue
        else:
            excess_str += f"{excess_name} is ${excess.get_text().strip()}\n"
    if not excess_str:
        return "No excess."
    return excess_str

def get_pol_type(file_name):
    '''
    Gets the type for the policy.
    '''
    if 'Hospital' in file_name:
        return 'Hospital'
    elif 'General' in file_name:
        return 'General'
    elif 'Combined' in file_name:
        return 'Combined'
    else:
        return "Can't find policy type."

def find_all_soup(soup, tag):
    '''
    Gets the text from an xml tag.
    '''
    found = soup.find_all(tag.lower())
    if not found:
        return ""
    return found[0].get_text().strip()

def get_wait(product):
    '''
    Get the waiting period 
    '''
    wait_str = ""
    wait_periods = product.find_all('waitingperiod')
    if not wait_periods:
        return "No waiting period found."
    for wait_period in wait_periods:
        try:
            title = str(wait_period['title'])
            unit  = str(wait_period['unit'])
            text  = wait_period.get_text().strip()
        except:
            wait_str = "Cannot find wait."
        wait_str += f"{title} wait is {text} {unit}s\n"
        
    return wait_str

def get_hosp_details(product, schema):
    '''
    Get all the hospital details required.
    '''
    covered = []
    not_covered = []
    limited_cover = []
    
    try:
        hosp_cover = product.find_all('hospitalcover')[0]
    except:
        return HospService("-", "-", "-", "-", "-", "-")
    cov_children = hosp_cover.findChildren()
    for child in cov_children:
        if "<medicalservices" in str(child): break
        tag = get_tag_name(str(child))
        info = child.get_text().strip()
        covered.append(f"{tag}: {info}")

    services = product.find_all('medicalservice')
    for service in services:
        if service['cover'] == "Covered":
            covered.append(str(service['title']))
        elif service['cover'] == "NotCovered":
            not_covered.append(str(service['title']))
        elif service['cover'] == "Restricted":
            limited_cover.append(str(service['title']))

    wait_element = product.find_all('waitingperiods')[0]
    wait_str = get_wait(wait_element)
    other    = find_all_soup(product, "otherproductfeatures")
    co_pay   = find_all_soup(product, 'copayments')
    if not co_pay:
        co_pay = "No copayment."

    hosp_service = HospService(covered, not_covered, limited_cover, wait_str, other, co_pay)

    return hosp_service

def get_benefits(service):
    '''
    Gets all the benefit elements for a service.
    '''
    benefit_elements = service.find_all('benefit')
    if not benefit_elements:
        return "No benefits found."
    ben_str = ""
    for benefit in benefit_elements:
        item = benefit['item'].strip()
        fee  = benefit.get_text().strip()
        typ  = benefit['type'].strip()
        ben_str += f"{item} {fee} {typ}\n"

    return ben_str

def get_general_services(product, schema):
    '''
    Gets all information about the general services.
    '''
    general_services = product.find_all('generalhealthservice')
    if not general_services:
        return "No general services found."
    
    general_details = {}
    for service in general_services:
        name = service['title'].strip()
        cover = service['covered'].strip()
        if cover == "false":
            serv_obj = GeneralService(name, cover, "-", "-", "-")
            general_details[name] = serv_obj
            continue
        wait_element = service.find_all('waitingperiod')[0]
        wait_str = f"{wait_element.get_text().strip()} {wait_element['unit']}s"
        benefit  = get_benefits(service)
        serv_obj = GeneralService(name, cover, wait_str, "", benefit)
        general_details[name] = serv_obj
        if schema == '2.0':
            general_details['Ambulance'] = GeneralService("Ambulance", "-", "-", "-", "-")

    get_benefit_limits(product, general_details)

    return general_details

def get_benefit_limits(product, general_details):
    '''
    Gets all the limits of general services.
    '''
    all_limits = product.find_all('benefitlimit')
    if not all_limits:
        return "No benefit limits found."
    for limit in all_limits:
        if limit['title'] == "Ambulance": continue
        children = limit.findChildren()
        fee = "No fee"
        for child in children:
            if 'limitper' in str(child):
                fee = child.get_text().strip()
        # Services with the same limit.
        services_with_limit = limit.find_all('service')
        for s in services_with_limit:
            service = s.get_text().strip()
            # Updates the limit in the general details dict.
            general_details[service].limits = fee
            if s['sublimitsapply'] == "true":
                general_details[service].limits += " sublimits apply"

def get_prov_arr(product, state, fund_obj, schema):
    '''
    Gets the provider arrangements for a policy.
    '''
    if schema == '3.0':
        provider_elements = product.find_all('productpreferredproviderservices')
        if not provider_elements:
            return "No provider arrangements found."
        
        provider_element = provider_elements[0]
        usefund = provider_element['usefund']
        # False: product contains the provider.
        if usefund == "false":
            return provider_element.get_text().strip()
        
        # True: provider can be found in the fund dictionary.
        provider_services = fund_obj.preferred_provider_services
        for provider in provider_services:
            if provider[0] == state:
                return provider[1]

        return "No provider arrangements found."
    elif schema == '2.0':
        provider_elements = product.find_all('preferredproviderservices')
        if not provider_elements:
            return "No provider arrangements found."
        return provider_elements[0].get_text().strip()

    return "Wrong schema."

def get_amb_info_new(product, fund_code, funds_dict):
    '''
    Gets ambulance information for schema 3 policies.
    '''
    product_ambulance = product.find_all('productambulance')
    if not product_ambulance:
        return "No ambulance information found."
    
    usefund = product_ambulance[0]['usefund']
    if usefund == "false":
        return product_ambulance[0].get_text().strip()

    fund_obj = funds_dict[fund_code]
    return fund_obj

def get_amb_info_old(product, general_dict):
    '''
    Gets ambulance information for schema 2 policies.
    '''
    if type(general_dict) is str:
        return 
    try:
        ambulance = product.find_all('generalhealthambulance')[0]
    except:
        print(type(general_dict))
        general_dict['Ambulance'].limits = "Cant find ambulance info."
        return
    cover = ambulance['cover']
    if cover == "Full" or cover == "Part":
        wait = ambulance.find_all('waitingperiod')[0]
        period = wait.get_text().strip()
        unit   = wait['unit']
        wait_str = f"{period} {unit}"
        # Get the limits.
        limits = product.find_all('benefitlimit')
        benefit_limit = ""
        for limit in limits:
            if limit['title'] == "Ambulance":
                benefit_limit = find_all_soup(limit, 'annuallimit')
        general_dict['Ambulance'].limits = benefit_limit
        general_dict['Ambulance'].wait   = wait_str

def get_amb_other(other_list, state):
    '''
    Get other ambulance features from the list of possibilities.
    Each state has different ambulance features. 
    Can be found in the fund xml file.
    '''
    amb_other = "Not found."
    for other in other_list:
        if state == other[0]:
            return other[1]
    return amb_other

def schema_3(product, xml_file, funds_dict, schema, pdf_link):
    '''
    Gets all the required information from new policies, 
    i.e. schema 3.0 policies.
    '''
    pol_name   = find_all_soup(product, "name")
    fund_name  = find_all_soup(product, "fundcode")
    prod_code  = product['productcode'].strip()
    pol_id     = fund_name + "/" + prod_code
    pdf_link   = DOWNLOAD_LINK + pol_id
    status     = find_all_soup(product, "productstatus")
    excess_str = get_excess(product)
    mo_prem    = find_all_soup(product, "premiumnorebate")
    state      = find_all_soup(product, "state")
    adults     = prod_code[len(prod_code)-2]
    dpndnts    = find_all_soup(product, 'scale')
    # CHANGE THIS
    pol_type   = get_pol_type(xml_file)
    hosp_det   = get_hosp_details(product, schema)
    general    = get_general_services(product, schema)
    medicare   = find_all_soup(product, 'medicarelevysurchargeexempt')
    issue_date = find_all_soup(product, 'dateissued')
    prov_arr   = get_prov_arr(product, state, funds_dict[fund_name], schema)
    other      = find_all_soup(product, 'otherservices')
    avail_for  = funds_dict[fund_name].restrictions
    amb = get_amb_info_new(product, fund_name, funds_dict)
    if isinstance(amb, Fund):
        amb_emer = amb.amb_emer
        callout_fee = amb.amb_call_out_fees
        amb_other = get_amb_other(amb.amb_other, state)
    else:
        amb_emer = callout_fee = amb_other = amb
    try:
        corp = product.find_all('corporate')[0]['atomic']
    except:
        "No corporate information found."
    try:
        age_disc = product.find_all('agebaseddiscount')[0]['available']
    except:
        age_disc = "Cannot find information on youth discount."
    try:
        accident_cov = product.find_all('hospitalcover')[0]['accidentcover']
    except:
        accident_cov = "No accident cover."
    try:
        trav_accom_ben = product.find_all('hospitalcover')[0]['traveloraccommodationsbenefit']
    except:
        trav_accom_ben = "No travel and accommodation benefits."


    new_pol = NewPolicy(schema, pol_name, fund_name, pdf_link, status, excess_str, mo_prem, state, adults, \
        dpndnts, "None avail", pol_type, corp, medicare, issue_date, avail_for, \
        prov_arr, hosp_det, general, other, age_disc, trav_accom_ben, pol_id, accident_cov, amb_emer, \
        callout_fee, amb_other)

    return new_pol

def schema_2(product, xml_file_name, funds_dict, schema, pdf_link):
    '''
    Gets all the information required for old policies, 
    i.e. schema 2.0 policies.
    '''
    pol_name   = find_all_soup(product, "name")
    fund_name  = find_all_soup(product, "fundcode")
    prod_code  = product['productcode'].strip()
    pol_id     = fund_name + "/" + prod_code
    pdf_link   = DOWNLOAD_LINK + pol_id
    status     = find_all_soup(product, "productstatus")
    excess_str = get_excess(product)
    mo_prem    = find_all_soup(product, "premiumnorebate")
    state      = find_all_soup(product, "state")
    adults     = prod_code[len(prod_code)-2]
    dpndnts    = find_all_soup(product, 'category')
    pol_type   = find_all_soup(product, 'producttype')
    corp       = product.find_all('corporate')[0]['atomic'] 
    hosp_det   = get_hosp_details(product, schema)
    medicare   = find_all_soup(product, 'medicarelevysurchargeexempt')
    issue_date = find_all_soup(product, 'dateissued')
    prov_arr   = get_prov_arr(product, state, funds_dict[fund_name], schema)
    other      = find_all_soup(product, 'otherservices')
    avail_for  = funds_dict[fund_name].restrictions
    if pol_type != 'Hospital':
        general = get_general_services(product, schema)
        get_amb_info_old(product, general)
    else:
        general = ""

    old_pol = OldPolicy(schema, pol_name, fund_name, pdf_link, status, excess_str, mo_prem, state, adults, \
        dpndnts, "No avail", pol_type, corp, medicare, issue_date, avail_for, prov_arr, hosp_det, \
        general, other)

    return old_pol

def create_excel(destination):
    '''
    Sets up Excel Category Sheet & Bold title given sheetname
    '''
    wb = Workbook()
    ws = wb.active
    colname = ["PDF Type", "Name", "Fund", "PDFLink", "Status", "Excess", \
            "Monthly Premium", "State", "Adults", "Scale (Adults + Dependants)", \
            "Availability", "Policy Type", "Corporate Product", \
            "Hospital Cover During Visit", "Hospital Services not Covered", \
            "Hospital Services Limited Cover", "Waiting periods", "Copayment", \
            "Other Hospital Cover Features", "General Dental - WP", \
            "General Dental - Limits", "General Dental - Max Benefits", \
            "Major Dental - WP", "Major Dental - Limits", "Major Dental - Max Benefits", \
            "Endodontic - WP", "Endodontic - Limits", "Endodontic - Max Benefits", \
            "Orthodontic - WP", "Orthodontic - Limits", "Orthodontic - Max Benefits", \
            "Optical - WP", "Optical - Limits", "Optical - Max Benefits", \
            "NonPBSPharmaceuticals - WP", "NonPBSPharmaceuticals - Limits", \
            "NonPBSPharmaceuticals - Max Benefits", "Physio - WP", "Physio - Limits", \
            "Physio - Max Benefits", "Chiropractic - WP", "Chiropractic - Limits", \
            "Chiropractic - Max Benefits", "Podiatry - WP", "Podiatry - Limits", \
            "Podiatry - Max Benefits", "Psychology - WP", "Psychology - Limits", \
            "Psychology - Max Benefits", "Acupuncture - WP", "Acupuncture - Limits", \
            "Acupuncture - Max Benefits", "Naturopathy - WP", "Naturopathy - Limits", \
            "Naturopathy - Max Benefits", "Massage - WP", "Massage - Limits", \
            "Massage - Max Benefits", "HearingAids - WP", "HearingAids - Limits", \
            "HearingAids - Max Benefits", "BloodGlucose Monitoring - WP", \
            "BloodGlucose Monitoring - Limits", "BloodGlucose Monitoring - Max Benefits", \
            "Ambulance - Emergency", "Ambulance - Call out fees", "Ambulance - other information", \
            "Other Treatment Cover Features", "Medicare Surcharge Levy", "Issue Date", \
            "Available for", "Provider Arrangements", "Youth discount", \
            "Travel and accommodation beneft", "Policy ID", "Accident cover"]
    #Creates Bold Column Titles
    for i in range(len(colname)):
        ws.cell(row = 1, column = i + 1).value = colname[i]
        ws.cell(row = 1, column = i + 1).font = Font(size = 14, bold = True)
    wb.save(filename=destination)

def write_policy(excel, policy, input_row, ws):
    '''
    Writes all the policy information to the excel sheet passed in as ws.
    '''
    # Input basic policy information.
    ws.cell(row=input_row, column=COL_POL_NAME).value = policy.pol_name
    ws.cell(row=input_row, column=COL_FUND_NAME).value = policy.fund_name
    ws.cell(row=input_row, column=COL_PDF_LINK).value = policy.pdf_link
    ws.cell(row=input_row, column=COL_STATUS).value = policy.status
    ws.cell(row=input_row, column=COL_EXCESS).value = policy.excess
    ws.cell(row=input_row, column=COL_MOPREM).value = policy.mo_prem
    ws.cell(row=input_row, column=COL_STATE).value = policy.state
    ws.cell(row=input_row, column=COL_ADULTS).value = policy.adults
    ws.cell(row=input_row, column=COL_DPNDNTS).value = policy.dpndnts
    ws.cell(row=input_row, column=COL_AVAIL).value = policy.avail
    ws.cell(row=input_row, column=COL_POL_TYPE).value = policy.pol_type
    ws.cell(row=input_row, column=COL_CORP).value = policy.corp
    ws.cell(row=input_row, column=COL_ISSUE_DATE).value = policy.issue_date
    ws.cell(row=input_row, column=COL_AVAIL_FOR).value = policy.avail_for
    ws.cell(row=input_row, column=COL_PROV_ARR).value = policy.prov_arr
    ws.cell(row=input_row, column=COL_OTHER).value = policy.other
    medicare = policy.medicare
    if medicare == "true":
        ws.cell(row=input_row, column=COL_MEDICARE).value = "Exempted"
    else:
        ws.cell(row=input_row, column=COL_MEDICARE).value = "Not exempted"
    if policy.schema == '2.0':
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "OLD"
        ws.cell(row=input_row, column=COL_AGE_DISC).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_TRAV_ACCOM_BEN).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_POL_ID).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_ACCIDENT_COV).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_AMBULANCE_EMER).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_AMBULANCE_FEE).value = "OLD PDF. Does not contain this."
        ws.cell(row=input_row, column=COL_AMBULANCE_OTHER).value = "OLD PDF. Does not contain this."
    else:
        ws.cell(row=input_row, column=COL_PDF_TYPE).value = "NEW"
        ws.cell(row=input_row, column=COL_AGE_DISC).value = policy.youth_disc
        ws.cell(row=input_row, column=COL_TRAV_ACCOM_BEN).value = policy.travel_accom_ben
        ws.cell(row=input_row, column=COL_POL_ID).value = policy.pol_id
        ws.cell(row=input_row, column=COL_ACCIDENT_COV).value = policy.accident_cover
        ws.cell(row=input_row, column=COL_AMBULANCE_EMER).value = policy.amb_emer
        ws.cell(row=input_row, column=COL_AMBULANCE_FEE).value = policy.amb_callout_fees
        ws.cell(row=input_row, column=COL_AMBULANCE_OTHER).value = policy.amb_other

    
    # Inputting hospital cover details.
    covered = ""
    for c in policy.hosp_cover.covered:
        covered += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_COVERED).value = covered
    not_covered = ""
    for c in policy.hosp_cover.not_covered:
        not_covered += f"{c}, "
    limited_cover = ""
    ws.cell(row=input_row, column=COL_HOSP_NOT_COVERED).value = not_covered
    for c in policy.hosp_cover.limited_cover:
        limited_cover += f"{c}, "
    ws.cell(row=input_row, column=COL_HOSP_LIMITED).value = limited_cover
    ws.cell(row=input_row, column=COL_WAIT_PERIODS).value = policy.hosp_cover.wait
    ws.cell(row=input_row, column=COL_COPAYMENT).value = policy.hosp_cover.co_pay
    ws.cell(row=input_row, column=COL_OTHER_HOSP).value = policy.hosp_cover.other

    # Inputting general details.
    for s in policy.gen_services:
        col = ''
        if 'DentalGeneral' in s:
            col = COL_GENERAL_DENTAL
        elif 'DentalMajor' in s:
            col = COL_MAJOR_DENTAL
        elif 'Endodontic' in s:
            col = COL_ENDODONTIC
        elif 'Orthodontic' in s:
            col = COL_ORTHODONTIC
        elif 'Optical' in s:
            col = COL_OPTICAL
        elif 'NonPBS' in s:
            col = COL_NONPSBPHARM
        elif 'Physio' in s:
            col = COL_PHYSIO
        elif 'Chiro' in s:
            col = COL_CHIRO
        elif 'Podiatry' in s:
            col = COL_PODIATRY
        elif 'Psychology' in s:
            col = COL_PSYCH
        elif 'Acupuncture' in s:
            col = COL_ACUPUNC
        elif 'Naturopathy' in s:
            col = COL_NATUR
        elif 'Massage' in s:
            col = COL_MASSAGE
        elif 'HearingAids' in s:
            col = COL_HEARING
        elif 'Glucose' in s:
            col = COL_BLOOD
        elif 'Ambulance' in s and policy.schema == '2.0':
            col = COL_AMBULANCE_WP
        
        if col != '':
            ws.cell(row=input_row, column=col).value = policy.gen_services[s].wait 
            ws.cell(row=input_row, column=col+1).value = policy.gen_services[s].limits 
            ws.cell(row=input_row, column=col+2).value = policy.gen_services[s].max_ben 

def main():
    print(TITLE)

    # Gets all information about the funds. 
    print("... Scraping funds XML file...\n")
    funds_dict = parse_funds_file()

    # Get names of all the files we need to scrape.
    entries = os.listdir(f"{sys.path[0]}/privatehealth-04-apr-2019")
    files = []
    for e in entries:
        if ".xml" in e and "Funds" not in e:
            files.append(e)

    print("... Scraping the policy XML files ...\n")
    for file in files:
        xml_file_name = f"{sys.path[0]}/privatehealth-04-apr-2019/{file}"
        xml_file = open(xml_file_name)
        print("... Creating soup ...\n")
        xml_soup = BeautifulSoup(xml_file, "lxml")

        # Key: link, Value: policy object
        print(f"-- Scraping {file} file --")
        policies = {}
        pdf_link = ""
        product_count = xml_soup.find_all("products")[0]['count']
        print(f"Product Count in file: {product_count}")
        products = xml_soup.find_all("product")
        n_product = 1
        for product in products:
            print(f"Scraping product {n_product} of {product_count} from xml file.", end='\r')
            n_product += 1
            schema = product['schemaversion']
            if schema == '3.0':
                policy = schema_3(product, xml_file_name, funds_dict, schema, pdf_link)
            elif schema == '2.0':
                policy = schema_2(product, xml_file_name, funds_dict, schema, pdf_link)
            policies[policy.pdf_link] = policy
        print('\n')
        
        excel_dest = f"{sys.path[0]}/results/{file[:len(file)-4]}" + \
            datetime.datetime.now().strftime("%d %B %Y at %H.%M") + ".xlsx"
        print(f"-- Inputting into {excel_dest} excel file --")
        create_excel(excel_dest)

        wb = load_workbook(excel_dest)
        ws = wb.active
        n_pol = 1
        for p in policies:
            print(f"Filling policy {n_pol} out of {product_count} in excel sheet.", end='\r')
            write_policy(excel_dest, policies[p], n_pol + 1, ws)
            n_pol += 1
        print("\n")
        wb.save(excel_dest)


if __name__ == "__main__":
    main()